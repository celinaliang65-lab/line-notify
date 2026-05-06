#!/usr/bin/env python3
"""
LINE 班群自動通知系統
每天早上 8:00 執行（台灣時間），依 Excel 日期推播對應班群
"""

import os
import sys
import requests
import pandas as pd
from datetime import datetime, date
import zoneinfo
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

EXCEL_PATH   = "data/LINE班群提醒排程.xlsx"
SHEET_NAME   = "提醒排程"
LINE_API_URL = "https://api.line.me/v2/bot/message/push"

COL_TYPE    = "提醒類別"
COL_NAME    = "班群名稱"
COL_GROUP   = "目標班群 ID"
COL_DATE    = "提醒日期"
COL_TITLE   = "訊息標題"
COL_CONTENT = "提醒內容細節"
COL_SENT    = "已發送"
COL_SENT_AT = "發送時間"

# 提醒類別 → badge 顏色
CATEGORY_STYLE = {
    "作業繳交": {"bg": "#E3F2FD", "text": "#1565C0"},
    "上台規範": {"bg": "#FFF3E0", "text": "#E65100"},
    "公告":     {"bg": "#E8F5E9", "text": "#1B5E20"},
}
DEFAULT_STYLE = {"bg": "#F5F5F5", "text": "#555555"}


def get_token() -> str:
    token = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
    if not token:
        print("❌ 未設定環境變數 LINE_CHANNEL_ACCESS_TOKEN")
        sys.exit(1)
    return token


def send_line_message(token: str, group_id: str, category: str,
                      title: str, content: str, notify_date: str) -> bool:
    # ✅ 驗證 Group ID 格式（C 開頭、33碼）
    if not group_id.startswith("C") or len(group_id) != 33:
        print(f"   ❌ Group ID 格式錯誤：[{group_id}]（應為 C 開頭共 33 碼）")
        print(f"      請至 webhook.site 查看 Body 內的 groupId 欄位")
        return False

    style    = CATEGORY_STYLE.get(category, DEFAULT_STYLE)
    alt_text = f"【{category}】{title}｜{content}"

    payload = {
        "to": group_id,
        "messages": [
            {
                "type": "flex",
                "altText": alt_text,
                "contents": {
                    "type": "bubble",
                    "body": {
                        "type": "box",
                        "layout": "vertical",
                        "paddingAll": "16px",
                        "spacing": "sm",
                        "contents": [
                            {
                                "type": "box",
                                "layout": "horizontal",
                                "alignItems": "center",
                                "contents": [
                                    {
                                        "type": "box",
                                        "layout": "vertical",
                                        "backgroundColor": style["bg"],
                                        "paddingAll": "4px",
                                        "cornerRadius": "4px",
                                        "flex": 0,
                                        "contents": [
                                            {
                                                "type": "text",
                                                "text": category if category else "通知",
                                                "size": "xs",
                                                "weight": "bold",
                                                "color": style["text"]
                                            }
                                        ]
                                    },
                                    {
                                        "type": "text",
                                        "text": title if title else "班群通知",
                                        "size": "sm",
                                        "weight": "bold",
                                        "color": "#111111",
                                        "flex": 1,
                                        "margin": "sm",
                                        "wrap": True
                                    },
                                    {
                                        "type": "text",
                                        "text": notify_date,
                                        "size": "xs",
                                        "color": "#999999",
                                        "align": "end",
                                        "flex": 0
                                    }
                                ]
                            },
                            {
                                "type": "separator",
                                "margin": "sm",
                                "color": "#EEEEEE"
                            },
                            {
                                "type": "text",
                                "text": content,
                                "size": "sm",
                                "color": "#333333",
                                "wrap": True,
                                "margin": "xs"
                            }
                        ]
                    }
                }
            }
        ]
    }

    try:
        resp = requests.post(
            LINE_API_URL,
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {token}"},
            json=payload,
            timeout=10
        )
        if resp.status_code == 200:
            return True
        print(f"   ⚠️  LINE API 錯誤 {resp.status_code}: {resp.text}")
        return False
    except Exception as e:
        print(f"   ⚠️  連線失敗: {e}")
        return False


def run():
    token   = get_token()
    today   = datetime.now(zoneinfo.ZoneInfo("Asia/Taipei")).date()
    now_str = datetime.now(zoneinfo.ZoneInfo("Asia/Taipei")).strftime("%Y/%m/%d %H:%M")

    print(f"📅 執行日期：{today}")

    # 第1列為標題，跳過第2列說明列，資料從第3列開始
    # 先不強制 dtype=str，讓日期欄自動解析
    df_raw = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, header=0, skiprows=[1])
    df_raw.columns = df_raw.columns.str.strip()
    raw_dates = pd.to_datetime(df_raw[COL_DATE], errors="coerce").dt.date
    df = df_raw.astype(str)
    df.columns = df_raw.columns
    df[COL_DATE] = raw_dates

    # Debug：印出每列狀態
    print(f"📋 共 {len(df)} 筆資料（不含說明列）")
    for i, row in df.iterrows():
        d = row.get(COL_DATE)
        sent = row.get(COL_SENT)
        print(f"   [{i}] 日期={d} (type={type(d).__name__}) | 已發送={repr(sent)} | 日期符合={d==today} | 發送為空={pd.isna(sent) or str(sent).strip()=='' or str(sent).strip().lower()=='nan'}")

    # 篩選：今天 且 已發送為空
    mask = (
        df[COL_DATE].notna() &
        (df[COL_DATE] == today) &
        (df[COL_SENT].isna() | (df[COL_SENT].str.strip() == "") | (df[COL_SENT].str.strip().str.lower() == "nan"))
    )
    targets = df[mask]

    if targets.empty:
        print("✅ 今天沒有待發送的提醒。")
        return

    print(f"📨 找到 {len(targets)} 筆待發送\n")

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 欄位名稱 → 欄號（從第1列標題）
    col_map = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }
    sent_col    = col_map.get(COL_SENT)
    sent_at_col = col_map.get(COL_SENT_AT)

    sent_fill = PatternFill("solid", start_color="C8E6C9")
    auto_font = Font(italic=True, color="558B2F", size=10, name="Arial")

    # df index 0 = Excel 第3列（標題列1 + 說明列2 + 資料從第3列）
    EXCEL_DATA_START = 3

    success = 0
    for df_idx, row in targets.iterrows():
        group_id    = str(row.get(COL_GROUP, "")).strip()
        category    = str(row.get(COL_TYPE, "")).strip()
        title       = str(row.get(COL_TITLE, "")).strip()
        content     = str(row.get(COL_CONTENT, "")).strip()
        name        = str(row.get(COL_NAME, "")).strip()
        notify_date = str(row.get(COL_DATE, "")).strip()

        if not group_id or not content:
            print(f"   ⚠️  索引 {df_idx} 缺少群組 ID 或內容，跳過")
            continue

        print(f"   → 【{name}】{group_id}")
        print(f"      類別：{category}｜標題：{title}")
        print(f"      內容：{content}")

        ok = send_line_message(token, group_id, category, title, content, notify_date)
        if ok:
            excel_row = EXCEL_DATA_START + df_idx
            for col, val in [(sent_col, "TRUE"), (sent_at_col, now_str)]:
                if col:
                    cell = ws.cell(row=excel_row, column=col, value=val)
                    cell.fill = sent_fill
                    cell.font = auto_font
            print("      ✅ 成功")
            success += 1
        else:
            print("      ❌ 失敗")

    if success > 0:
        wb.save(EXCEL_PATH)
        print(f"\n🎉 完成！{success}/{len(targets)} 筆發送成功，Excel 已更新")
    else:
        print(f"\n⚠️  完成！{success}/{len(targets)} 筆發送成功，Excel 不回寫")


if __name__ == "__main__":
    run()
