#!/usr/bin/env python3
"""
fetch_groups.py — 執行一次即可
讀取「班群設定」工作表中的 LINE Group ID，
呼叫 LINE API 查詢群組名稱，自動填回 Excel。
"""

import os
import sys
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

EXCEL_PATH  = "data/LINE班群提醒排程.xlsx"
SHEET_NAME  = "班群設定"
LINE_API    = "https://api.line.me/v2/bot/group/{group_id}/summary"

COL_NAME    = "班群名稱"
COL_ID      = "LINE Group ID"
COL_NOTE    = "備註"


def get_token() -> str:
    token = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
    if not token:
        # 允許直接從命令列傳入：python fetch_groups.py <token>
        if len(sys.argv) > 1:
            return sys.argv[1]
        print("❌ 請設定環境變數 LINE_CHANNEL_ACCESS_TOKEN")
        print("   或執行：python fetch_groups.py <your_token>")
        sys.exit(1)
    return token


def fetch_group_name(token: str, group_id: str) -> str | None:
    """呼叫 LINE API 取得群組名稱，失敗回傳 None"""
    url = LINE_API.format(group_id=group_id.strip())
    headers = {"Authorization": f"Bearer {token}"}
    try:
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code == 200:
            return resp.json().get("groupName", "")
        elif resp.status_code == 403:
            print(f"   ⚠️  Bot 尚未加入群組 {group_id}，請先將 Bot 加為群組成員")
        elif resp.status_code == 404:
            print(f"   ⚠️  找不到群組 {group_id}，請確認 Group ID 正確")
        else:
            print(f"   ⚠️  API 錯誤 {resp.status_code}: {resp.text}")
        return None
    except Exception as e:
        print(f"   ⚠️  連線失敗: {e}")
        return None


def run():
    token = get_token()

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 建立欄位名稱 → 欄號對照
    col_map = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }
    name_col = col_map.get(COL_NAME)
    id_col   = col_map.get(COL_ID)

    if not name_col or not id_col:
        print(f"❌ 找不到欄位「{COL_NAME}」或「{COL_ID}」，請確認工作表格式")
        sys.exit(1)

    filled_font  = Font(name="Arial", size=10, color="1B5E20", bold=True)
    filled_fill  = PatternFill("solid", start_color="C8E6C9")
    skip_font    = Font(name="Arial", size=10, color="9E9E9E", italic=True)

    updated = 0
    skipped = 0

    print(f"🔍 開始查詢群組名稱...\n")

    for row in range(2, ws.max_row + 1):
        group_id = ws.cell(row=row, column=id_col).value
        existing = ws.cell(row=row, column=name_col).value

        # 跳過空白列
        if not group_id or str(group_id).strip() == "":
            continue

        group_id = str(group_id).strip()

        # 已有名稱就跳過，不覆蓋
        if existing and str(existing).strip() not in ("", "班群名稱"):
            print(f"   ⏭  {group_id} → 已有名稱「{existing}」，跳過")
            skipped += 1
            continue

        print(f"   🔎 查詢 {group_id} ...", end=" ")
        name = fetch_group_name(token, group_id)

        if name:
            cell = ws.cell(row=row, column=name_col, value=name)
            cell.font = filled_font
            cell.fill = filled_fill
            print(f"✅ {name}")
            updated += 1
        else:
            cell = ws.cell(row=row, column=name_col, value="（查詢失敗）")
            cell.font = skip_font
            print("❌ 查詢失敗")

    wb.save(EXCEL_PATH)
    print(f"\n🎉 完成！更新 {updated} 筆，略過 {skipped} 筆")
    print(f"   已儲存至 {EXCEL_PATH}")


if __name__ == "__main__":
    run()
